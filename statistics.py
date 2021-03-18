from openpyxl import *
import copy
from numpy import * 
import datastructure as ds
import numpy as np


# 
def statistics(filename, A, x, b,  numf, numg, fb2col, FP, famsize, bundle2rank):
	print("+++++++++++++++++++++++++++++++++++ Statistics +++++++++++++++++++++++++++++++++++++")
	print(FP)
	
	
	nmf, nmp, bysize, fbyng, pbyng, fbypref, pbypref = stathelper(x, numf, numg, fb2col, FP, famsize)
	
	print("nmf = " + str(nmf))
	print("nmp = " + str(nmp))
	print(bysize)
	print(fbyng)
	print(pbyng)
	print(fbypref)	
	print(pbypref)
	
	avgnmf = mean(nmf)
	avgnmp = mean(nmp)
	avgbysize = [0] * 5
	
	# bundle rank
	brank, bwrank, avg, avgwrank, count, countbyp, match = matchbundlerank(x, numf, numg, fb2col, bundle2rank, famsize)
	print('brank =' + str(brank))
	print('count =' + str(count))
	print(match)
	rbysize = rankbysize(brank, famsize) 
	
	# Violations
	V, P = violations(A,x,b)
	
	maxv = max(P)
	print('V = ' + str(V))
	print('P = ' + str(P))
	print('max violation =' + str(maxv))
	
	#
	for i in range(5):
		avgbysize[i] = mean(bysize[i])
		
	# Envy
	envy, numenvy, wenvy = countenvy(brank, famsize)
	print('envy =' + str(envy))
	print(numenvy)
	print('wenvy = ' + str(wenvy))
	
	aenvybysize, anumenvybysize, awenvybysize = countenvybysize(envy, numenvy, wenvy, famsize)
	
	avgenvy = sum(envy)/len(envy)
	avgnumenvy = sum(numenvy)/len(numenvy)
	avgwenvy = sum(wenvy)/len(wenvy)
	
	print('aenvybysize = ' + str(aenvybysize))
	
	print('Average envy = ' + str(avgenvy))
	print('Average number of envy families = ' + str(avgnumenvy))
	print('Average weighted envy = ' + str(avgwenvy))
	
	## Save to a file 
	wb=load_workbook(filename)
	
	# save the matching 
	ws=wb["Sheet1"]
	for i in range(numf):
		for j in range(numg):
			wcell = ws.cell(3+i, 39+j)
			wcell.value = round(match[i,j])
	
	
	ws=wb["Sheet3"]
	
	for i in range(numg):
		wcell = ws.cell(3, 35+i)
		wcell.value = nmf[i]
		
	for i in range(numg):
		wcell = ws.cell(4, 35+i)
		wcell.value = nmp[i]
		
	for i in range(5):
		for j in range(numg):
			wcell = ws.cell(6+i, 35+j)
			wcell.value = bysize[i][j]
	
	
	wcell = ws.cell(13, 10)
	wcell.value = 'Scarf'
	wcell = ws.cell(14, 10)
	wcell.value = avgnmf
	wcell = ws.cell(15, 10)
	wcell.value = avgnmp
	for i in range(5):
		wcell = ws.cell(16+i, 10)
		wcell.value = avgbysize[i]
		
	for i in range(numg):
		wcell = ws.cell(23+i, 10)
		wcell.value = fbyng[i]
	
	for i in range(numg):
		wcell = ws.cell(32+i, 10)
		wcell.value = pbyng[i]
		
	ws=wb["Sheet4"]
	for i in range(numg):
		wcell = ws.cell(19+i, 10)
		wcell.value = fbypref[i]
		
	for i in range(numg):
		wcell = ws.cell(28+i, 10)
		wcell.value = pbypref[i]
		
	# Save bundlerank
	ws=wb["Sheet5"]
	numb = len(bundle2rank[0])
	print('numb = ' + str(numb))
	for i in range(numb+1):
		wcell = ws.cell(2+i, 7)
		wcell.value = count[i]
		
	for i in range(numb+1):
		wcell = ws.cell(2+i, 16)
		wcell.value = countbyp[i]
		
	ws = wb["Sheet6"]
	for i in range(numf):
		wcell = ws.cell(2+i, 7)
		wcell.value = brank[i]
		
	ws=wb["Sheet7"]
	wcell = ws.cell(2, 7)
	wcell.value = avg
	
	wcell = ws.cell(5, 7)
	wcell.value = round(std(brank),2)
	
	wcell = ws.cell(8, 7)
	wcell.value = avgwrank
	
	wcell = ws.cell(11, 7)
	wcell.value = round(std(bwrank),2)
	
	for i in range(5):
		wcell = ws.cell(14+i, 7)
		wcell.value = round(rbysize[i],2)
	
	# Save envy
	ws=wb["Sheet8"]
	wcell = ws.cell(2, 7)
	wcell.value = avgenvy
	
	wcell = ws.cell(5, 7)
	wcell.value = avgnumenvy
	
	wcell = ws.cell(8, 7)
	wcell.value = avgwenvy
	
	
	for i in range(5):
		wcell = ws.cell(11+i, 7)
		wcell.value = round(aenvybysize[i],2)

	for i in range(5):
		wcell = ws.cell(11+i, 15)
		wcell.value = round(anumenvybysize[i],2)	
		
	for i in range(5):
		wcell = ws.cell(11+i, 23)
		wcell.value = round(awenvybysize[i],2)
	
	
	
	wb.save(filename)
		
	return 

	
# Helper function 
def stathelper(x, numf, numg, fb2col, FP, famsize):
	# Intialize
	nmf = [0] * numg
	nmp = [0] * numg
	fbyng = [0] * numg
	pbyng = [0] * numg
	fbypref = [0] * numg
	pbypref = [0] * numg
	
	bysize = zeros((5,numg), dtype=int32)
	
	# Compute
	for i in range(len(x)):
		if x[i] >= 1: 
			(f, b) = ind2fb(i, fb2col, numf + numg)
			size = getbundlesize(b)
			fbyng[size-1] += x[i]
			pbyng[size-1] += x[i] * famsize[f]
			
			#print(b)
			#print(bysize)
			for j in range(numg):
				if b[j] >= 1:
					nmf[j] += x[i]
					nmp[j] += x[i] * famsize[f]
					
					bysize[famsize[f]-1][j] += x[i] 
					#print('here')
					##print(j)
					#print(bysize)
					
					fbypref[FP[f][j] - 1] += x[i]
					pbypref[FP[f][j] - 1] += x[i] * famsize[f]
					
	return nmf, nmp, bysize, fbyng, pbyng, fbypref, pbypref
	
# Compute violations
def violations(A, x, b):
	realb = mul(A,x)
	V = [0] * len(b)
	P = [0] * len(b)
	
	for i in range(len(realb)):
		if (realb[i] > b[i]):
			V[i] = realb[i] - b[i]
			P[i] = round(V[i]/b[i] * 100, 1)
	
	print(realb)
	return V, P
	
			
def ind2fb(ind, fb2col, numrows):
	#print('ind = ' + str(ind))
	# offset by the number of rows 
	for key, value in fb2col.items():
		if value == (ind + numrows):
			#print('key = '  + str(key))
			return key
			
			
def getbundlesize(bundle):
	size = 0
	for i in bundle:
		if i > 0:
			size += 1

	return size
	
def mean(x):
	return round(sum(x)/len(x))

	
# Multiply a matrix with a vector
def mul(A, x):
	return [ds.dotproduct(a,x) for a in A]
	
# Get bundle rank from the match
def matchbundlerank(x, numf, numg, fb2col, bundlerank, famsize):
	numb = len(bundlerank[0])
	brank = [numb+1] * numf 
	bwrank = [numb+1] * sum(famsize)
	count = [0] * (numb+1)
	countbyp = [0] *(numb+1)
	s = 0
	sbysize = 0
	num = 0
	match = np.zeros((numf, numg))
	ind = 0
	
	for i in range(len(x)):
		if x[i] >= 1- 10**(-6):
			print(x[i])
			(f, b) = ind2fb(i, fb2col, numf + numg)
			
			for key, value in bundlerank[f].items():
				if key == b:
					brank[f] = value + 1 # offset by 1 
					match[f, :] = list(b)
					
					for j in range(famsize[f]):
						bwrank[ind] = brank[f]
						ind += 1
					
					# count the number of families/people get i-th bundle
					count[value] += x[i]
					countbyp[value] += x[i] * famsize[f]
					
					s += (value + 1) * x[i]
					sbysize += (value + 1) * x[i] * famsize[f]
					num += x[i]
					
	# count number of unmatched families				
	# count[numb] = numf - sum(count[0:numb])				
					
	# avgerage rank of matched families
	avg = round(s/num, 1)
	
	avgbysize = round(sbysize/sum(famsize))
			
	return brank, bwrank, avg, avgbysize, count, countbyp, match
	
## Count enviness
def countenvy(brank, S):	
	nF = len(S)
	envy = [0] * nF
	numenvy = [0] * nF
	wenvy = [0] * nF
	
	for f in range(nF):
		sum = 0
		count = 0
		maxdiff = 0
		for h in range(nF):
			
			if (f != h) and (S[f] == S[h]) and (brank[f] > brank[h]): # envy
				currdiff = brank[f] - brank[h]
				sum += currdiff
				
				numenvy[f] = numenvy[f]+1
				
				if maxdiff < currdiff:
					maxdiff = currdiff
			if (S[f] == S[h]):
				count += 1
		envy[f] = maxdiff
		if count > 1:
			wenvy[f] = sum/(count -1)
		
	return envy, numenvy, wenvy
	
def rankbysize(brank, S):
	nF = len(S)
	numfbysize = [0] * 5
	for f in range(nF):
		numfbysize[S[f]-1] += 1
		
	count = [0] * 5
	stari = 0
	
	for i in range(5):
		endi = stari + numfbysize[i]
		count[i] = average(brank[stari:endi])
		#print('here')
		#print(count[i])
		#print('sub envy = ' + envy )
		stari = endi
		
	return count


	
def countenvybysizehelper(envy, numfbysize):
	count = [0] * 5
	stari = 0
	
	for i in range(5):
		endi = stari + numfbysize[i]
		count[i] = average(envy[stari:endi])
		#print('here')
		#print(count[i])
		#print('sub envy = ' + envy )
		stari = endi
		
	return count


def countenvybysize(envy, numenvy, wenvy, S):
	nF = len(S)
	numfbysize = [0] * 5
	for f in range(nF):
		numfbysize[S[f]-1] += 1
	
	#print('numfbysize =' + str(numfbysize))
		
	aenvybysize = countenvybysizehelper(envy, numfbysize)
	anumenvybysize = countenvybysizehelper(numenvy, numfbysize)
	awenvybysize = countenvybysizehelper(wenvy, numfbysize)
	
	return aenvybysize, anumenvybysize, awenvybysize
		
# compute average (mean) of a list of numbers		
def average(alist):
	l = len(alist)
	ans = 0
	if l > 0:
		ans = sum(alist)/l
		
	return ans
		
	
		
	


	

	