import openpyxl as op
import sys
import csv

#argv[1]: input file name
#argv[2]: output file name
#argv[3]: budget for seniors, each non-senior has 1 dollar
#argv[4]: offset for alpha, i.e. extra number of seats needed
#argv[5]: capacity of a game
#argv[6]: upper bound of bundle size

book = op.load_workbook(sys.argv[1])
sheet = book.get_sheet_by_name("Sheet2")
row_num = sheet.max_row
col_num = sheet.max_column
gnum = 0
for j in range(1,col_num):
    if sheet.cell(row=2,column=j).value:
        gnum += 1
    else:
        break

plist = []
#plist[f][i] denotes the i-th favorite bundle of family f
blist = []
#blist[f] denotes the budget of family f
extra = int(float(sys.argv[4]))

sbud = float(sys.argv[3])

ub = int(float(sys.argv[6]))

for i in range(2,row_num+1):
    inplist = []
    rank = []
    for j in range(1,gnum+1):
        rank.append(sheet.cell(row=i,column=j).value)
    fsize = sheet.cell(row=i,column=gnum+2).value
    snum = sheet.cell(row=i,column=gnum+4).value
    for j in range(pow(2,gnum)-1, 0, -1):
        bundle = [0]*gnum
        bilist = [int(i) for i in list('{0:0b}'.format(j))]
        if sum(bilist) <= ub:
            bisize = len(bilist)
            for k in range(0,gnum-bisize):
                bilist.insert(0,0)
            for k in range(0,gnum):
                if bilist[k]==1:
                    bundle[rank[k]-1] = fsize + extra
            inplist.append(bundle)
    plist.append(inplist)
    blist.append(fsize-snum + snum*sbud)

with open(sys.argv[2], 'w', newline='') as csvfile:
    csvwriter = csv.writer(csvfile)
    csvwriter.writerow([row_num-1, gnum])
    for i in range(0,row_num-1):
        row = []
        row.append(blist[i])
        for j in range(0,len(plist[i])):
            #print(plist[j])
            #print(','.join(map(str,plist[j])))
            row.append(','.join(map(str,plist[i][j])))
        csvwriter.writerow(row)
    csvwriter.writerow([int(float(sys.argv[5]))]*gnum)
