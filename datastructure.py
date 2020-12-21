import collections
import csv
import numpy
import sys
import openpyxl as op

#################### MAIN ###########################


#numF = 2;

#s1 = list(1, 2, 4)
#s2 = list(3, 5, 6)
#p1 = list(1, 1, 1)
#p2 = list(0, 0, 0)

#c111 = tuple(1, s1, p1)
#c211 = tuple(2, s1, p1)
#c121 = tuple(1, s2, p1)
#c221 = tuple(2, s2, p1)





#####################################################

# @row: a row that can be either a family or a game
# @a: the first contract
# @b: the second contract
# returns YES if the first is more preferred than the second
# Otherwise returns NO

def strictlyprefer(a, b, row, numf, fp):
	if (row < numf):
		return fstrictlyprefer(a, b, row, numf, fp)
	else:
		return gstrictlyprefer(a, b, row, numf)

# for game row
# g is offset by  + numF
def gstrictlyprefer(a, b, g, numf):
	tol = 10**(-6)
	sa = isslack(a)
	sb = isslack(b)

	# Check if they are equal
	if (a == b):
		return False
	if (a[0] == -g-1):		# a is an active slack variable
		return False

	elif (b[0] == -g-1):		# b is an active slack variable
		return True
	# both are not active
	if (sa and sb):
		return (a[0] > b[0])
	elif (sa and (not sb)):
		return True
	elif ((not sa) and sb):
		return False
	else:					# both are non-slack variable
		za = iszerocoeff(a, g, numf)
		zb = iszerocoeff(b, g, numf)

		if (za and zb):
			return breaktie(a,b)
		elif (za and (not zb)):
			return True
		elif ((not za) and zb):
			return False
		else:				# both are non-zeros
			g = g - numf
			if not isequal(a[2][g], b[2][g]):
				if (a[2][g] + tol > b[2][g]):	# compare price
					return True
				elif (a[2][g] < b[2][g] + tol):
					return False
			else: 		# break tie
				return breaktie(a,b)

# for a family row
def fstrictlyprefer(a, b, f, numf, fp):
	#print("calling fstrictlyprefer")
	tol = 10**(-6)
	sa = isslack(a)
	sb = isslack(b)

	# Check if they are equal
	if (a == b):
		return False
	if (a[0] == -f-1):		# a is an active slack variable
		return False

	elif (b[0] == -f-1):		# b is an active slack variable
		return True
	# both are not active
	if (sa and sb):
		return (a[0] > b[0])
	elif (sa and (not sb)):
		return True
	elif ((not sa) and sb):
		return False
	else:					# both are non-slack variable
		za = iszerocoeff(a, f, numf)
		zb = iszerocoeff(b, f, numf)

		if (za and zb):
			#print("Both zeros")
			#print (breaktie(a,b))
			return breaktie(a,b)
		elif (za and (not zb)):
			return True
		elif ((not za) and zb):
			return False
		else:				# both are non-zeros
			#print("--------- Both non-zeros")
			#print(a)
			#print(b)
			#print(f)
			#print(fp)
			if (fp[f][a[1]] < fp[f][b[1]]):
				return True
			elif (fp[f][a[1]] > fp[f][b[1]]):
				return False
			else:
				msa = dotproduct(a[1], a[2])	# money spent
				msb = dotproduct(b[1], b[2])
				if not isequal(msa, msb):
					if (msa + tol > msb):
						return False
					elif (msa < msb + tol):
						return True
				else:
					return breaktie(a,b)

#
def weaklyprefer(a,b,row,numf, fp):
	if isequalcon(a,b):
		return True
	else:
		return strictlyprefer(a,b,row, numf,fp)


# dot product of two vectors with the same length
def dotproduct(x, y):
	sum = 0
	for i in range(len(x)):
		sum = sum + x[i] * y[i]

	return sum

# Break tie two contracts a and b
def breaktie(a,b):
	if (a[0] < b[0]):
		return True
	elif (a[0] > b[0]):
		return False
	else:
		if (a[1] == b[1]):
			return breaktievector(a[2], b[2])
		else:
			return breaktievector(a[1], b[1])

# @a: the first bundle
# @b: the second bundle
def breaktievector(a, b):
	tol = 10**(-6)
	n = len(a)
	for i in range(n):
		if not isequal(a[i], b[i]):
			if (a[i] < b[i] + tol):
				return True
			elif (a[i] + tol > b[i]):
				return False

	#print("+++++++ Break tie vector: SAME ++++++")
	return False

#
def isslack(c):
	return (c[0] < 0)

# for non-slack only
def iszerocoeff(a, row, numf):
	if (row < numf):
		return not(a[0] == row)
	else:
		return a[1][row - numf] == 0


# check if the same floating point values
def isequal(a, b):
	tol = 10**(-6)
	return abs(a-b) <= tol

#
def isequalprice(a, b):
	#print(a)
	#print(b)
	for i in range(len(a)):
		if not isequal(a[i], b[i]):
			return False

	#print("TRUEEEEEEEEEEEEEEEEEEEEEEE")
	return True

#
def isequalcon(c, d):
	#print("c =" +str(c))
	#print("d =" +str(d))
	return (c[0] == d[0]) and (c[1] == d[1]) and isequalprice(c[2], d[2])

#
def init(str):

    bundle2rank = [] #bundle maps to the rank, each family has one dictionary
    bundlelist = [] #preference list over bundles, each family has one list
    sortedbundle = [] #bundle of interest in incearsing alphabetic order, each family has one
    fb2col = {} #map (family,bundle) to the column index of matrix A
    numcol = 0

    #initialization and create slack contracts
    with open(str) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        line_count = 0
        for row in csv_reader:
            if line_count == 0:
                item_count = 0
                for item in row:
                    if item_count == 0:
                        numF = int(float(item)) #get number of families
                        budget = [0]*numF #create budget list
                        item_count = 1
                    elif item_count == 1:
                        numG = int(float(item)) #get number of games
                        break
                line_count += 1
            elif line_count < numF+1:
                item_count = 0
                for item in row:
                    if item_count == 0:
                        budget[line_count-1] = float(item) #budget[f-1] denotes the budget of family f
                        item_count = 1
                        bundle2rank.append({})
                        bundlelist.append([])
                        unsortedlist = []
                    elif item_count > 0:
                        if item.strip(): #filter empty string
                            intlist = [int(float(i)) for i in item.split(',')] #convert string to int
                            unsortedlist.append(intlist)
                            inttuple = tuple(intlist)
                            bundle2rank[line_count-1][inttuple] = item_count #bundle2rank[f-1] maps from a tuple bundle to rank for family f
                            bundlelist[line_count-1].append(inttuple)
                            #fb2col[(line_count-1,inttuple)] = numF + numG + numcol # (f-1, bundle) maps to the column index of A
                            item_count += 1
                            #numcol += 1
            #print(bundle2rank[line_count-1])
            #print(bundlelist[line_count-1])
                unsortedlist.sort()
                sortedbundle.append(unsortedlist)
                for item in unsortedlist:
                    inttuple = tuple(item)
                    fb2col[(line_count-1,inttuple)] = numF + numG + numcol
                    numcol += 1
                line_count += 1
            else:
                capacity = [int(float(i)) for i in row if i.strip()] #capacity[g-1] is the capacity of game g
                #print(capacity)

#now we have bundle2rank, bundlerank, fb2col, capacity, and budget
#print(fb2col)

    A = numpy.zeros((numF+numG,numF+numG+numcol))

#create slack columns/Contracts
#in fb2col, negative family/game id only has the id as the key, no bundle needed
#clist = [] #contract list
    for i in range(numF):
#    clist.append((-1*(i+1),[],[]))
        fb2col[(-1*(i+1), ())] = i
        A[i,i] = 1

    for i in range(numG):
#    clist.append((-1*(i+1+numF),[],[]))
        fb2col[(-1*(i+1+numF), ())] = i+numF
        A[i+numF,i+numF] = 1

    col_count = numF + numG
    for i in range(numF):
        for j in sortedbundle[i]:
            A[i, col_count] = 1
            game_count = 0
            for k in j:
                A[numF + game_count, col_count] = k
                game_count += 1
            col_count += 1

    return numF, numG, bundle2rank, bundlelist, fb2col, budget, capacity, numcol, A
#print(A)

def init_v2(filename,sbud,extra,cap,ub):

    book = op.load_workbook(filename)
    sheet = book.get_sheet_by_name("Sheet2")
    row_num = sheet.max_row
    col_num = sheet.max_column
    gnum = 0
    fnum = row_num-1
    for j in range(1,col_num):
        if sheet.cell(row=2,column=j).value:
            gnum += 1
        else:
            break

    bundle2rank = [] #bundle maps to the rank, each family has one dictionary
    bundlelist = [] #preference list over bundles, each family has one list
    sortedbundle = [] #bundle of interest in incearsing alphabetic order, each family has one
    fb2col = {} #map (family,bundle) to the column index of matrix A
    pglist = [] #plist[f][j] denotes family f's j-th most favorite game
    blist = [] #blist[f-1] denotes the budget of family f
    famsize = [] #famsize[f] denotes the size of family f
    numcol = 0
    b = []

    for i in range(2,row_num+1): #family i-2
        rank = []
        bundle2rank.append({})
        bundlelist.append([])
        unsortedlist = []
        for j in range(1,gnum+1):
            rank.append(sheet.cell(row=i,column=j).value)
        pglist.append(rank)
        fsize = sheet.cell(row=i,column=gnum+2).value
        snum = sheet.cell(row=i,column=gnum+4).value
        gsize = sheet.cell(row=i,column=gnum+6).value
        b.append(gsize)
        famsize.append(fsize)
        item_count = 0
        for j in range(pow(2,gnum)-1, 0, -1):
            bundle = [0]*gnum
            bilist = [int(k) for k in list('{0:0b}'.format(j))]
            if sum(bilist) <= ub:
                bisize = len(bilist)
                for k in range(0,gnum-bisize):
                    bilist.insert(0,0)
                for k in range(0,gnum):
                    if bilist[k]==1:
                        bundle[rank[k]-1] = fsize + extra
                unsortedlist.append(bundle)
                inttuple = tuple(bundle)
                bundle2rank[i-2][inttuple] = item_count+1 #bundle2rank[f-1] maps from a tuple bundle to rank for family f
                bundlelist[i-2].append(inttuple)
                item_count += 1
        blist.append(fsize-snum + snum*sbud)
        unsortedlist.sort()
        sortedbundle.append(unsortedlist)
        for item in unsortedlist:
            inttuple = tuple(item)
            fb2col[(i-2,inttuple)] = fnum + gnum + numcol
            numcol += 1
    b = b + [cap]*gnum
    #initialization and create slack contracts
    numF = fnum
    numG = gnum
    A = numpy.zeros((numF+numG,numF+numG+numcol))
    for i in range(numF):
#    clist.append((-1*(i+1),[],[]))
        fb2col[(-1*(i+1), ())] = i
        A[i,i] = 1

    for i in range(numG):
#    clist.append((-1*(i+1+numF),[],[]))
        fb2col[(-1*(i+1+numF), ())] = i+numF
        A[i+numF,i+numF] = 1

    col_count = numF + numG
    for i in range(numF):
        for j in sortedbundle[i]:
            A[i, col_count] = 1
            game_count = 0
            for k in j:
                A[numF + game_count, col_count] = k
                game_count += 1
            col_count += 1

    return fnum, gnum, bundle2rank, bundlelist, fb2col, blist, numcol, A, b, pglist, famsize

def init_v3(filename,sbud,cap):

    book = op.load_workbook(filename)
    sheet = book.get_sheet_by_name("Sheet2")
    row_num = sheet.max_row
    col_num = sheet.max_column
    gnum = 0
    fnum = row_num-1
    for j in range(1,col_num):
        if sheet.cell(row=2,column=j).value:
            gnum += 1
        else:
            break

    bundle2rank = [] #bundle maps to the rank, each family has one dictionary
    bundlelist = [] #preference list over bundles, each family has one list
    sortedbundle = [] #bundle of interest in incearsing alphabetic order, each family has one
    fb2col = {} #map (family,bundle) to the column index of matrix A
    pglist = [] #plist[f][j] denotes family f's j-th most favorite game
    blist = [] #blist[f-1] denotes the budget of family f
    famsize = [] #famsize[f] denotes the size of family f
    numcol = 0
    b = []

    for i in range(2,row_num+1): #family i-2
        bundle2rank.append({})
        bundlelist.append([])
        unsortedlist = []
        fsize = sheet.cell(row=i,column=gnum+2).value
        snum = sheet.cell(row=i,column=gnum+4).value
        gsize = sheet.cell(row=i,column=gnum+6).value
        b.append(gsize)
        famsize.append(fsize)
        item_count = 0
        for j in range(gnum+8,col_num):
            item = sheet.cell(row=i,column=j).value
            intlist = [int(float(i)) for i in item.split(',')] #convert string to int
            unsortedlist.append(intlist)
            inttuple = tuple(intlist)
            bundle2rank[i-2][inttuple] = item_count #bundle2rank[f-1] maps from a tuple bundle to rank for family f
            bundlelist[i-2].append(inttuple)
            item_count += 1
        blist.append(fsize-snum + snum*sbud)
        unsortedlist.sort()
        sortedbundle.append(unsortedlist)
        for item in unsortedlist:
            inttuple = tuple(item)
            fb2col[(i-2,inttuple)] = fnum + gnum + numcol
            numcol += 1
    b = b + [cap]*gnum
    #initialization and create slack contracts
    numF = fnum
    numG = gnum
    A = numpy.zeros((numF+numG,numF+numG+numcol))
    for i in range(numF):
#    clist.append((-1*(i+1),[],[]))
        fb2col[(-1*(i+1), ())] = i
        A[i,i] = 1

    for i in range(numG):
#    clist.append((-1*(i+1+numF),[],[]))
        fb2col[(-1*(i+1+numF), ())] = i+numF
        A[i+numF,i+numF] = 1

    col_count = numF + numG
    for i in range(numF):
        for j in sortedbundle[i]:
            A[i, col_count] = 1
            game_count = 0
            for k in j:
                A[numF + game_count, col_count] = k
                game_count += 1
            col_count += 1

    return fnum, gnum, bundle2rank, bundlelist, fb2col, blist, numcol, A, b, pglist, famsize

#generate ordlist where ordlist[i] is the preference of row i over the column index w.r.t the C matrix (i is 0-based)
def genordlist(A, numf, fp, bundlelist, fb2col):
	ordlist = []
	for i in range(len(A)):
		ordlist.append([])
		#get type 1 columns
		for j in range(len(A)):
			if i != j:
				ordlist[i].append(j)

		#get type 2 columns
		for j in range(len(A),len(A[i])):
			if A[i][j]==0:
				ordlist[i].append(j)

		#get type 3 columns
		if i < numf: #i is a family row
			for j in (bundlelist[i]):
				ordlist[i].append(fb2col[(i,j)])
		else: #i is a game row
			for j in range(len(A),len(A[i])):
				if A[i][j] > 0:
					ordlist[i].append(j)

		#get type 4 column
		ordlist[i].append(i)
	return ordlist

# contract to fb
def contract2fb(c):
	return (c[0], c[1])

# Function for printing out row minimizers
def printbasis(basis, fb2col):
	for i in range(len(basis)):
		c = basis[i]
		print(str(c) + " : " + str(fb2col[(c[0], c[1])]))
